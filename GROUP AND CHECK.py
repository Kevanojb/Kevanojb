import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from tkinter import simpledialog

def have_played_together(player1, player2, past_games_data, max_games):
    games_to_check = sorted(list(past_games_data.keys()), reverse=True)[:max_games]
    for game_key in games_to_check:
        game_df = past_games_data[game_key]
        if 'GROUP NUMBER' in game_df.columns and 'PLAYER NAME' in game_df.columns:
            for group_number in game_df['GROUP NUMBER'].unique():
                group_df = game_df[game_df['GROUP NUMBER'] == group_number]
                if player1 in group_df['PLAYER NAME'].values and player2 in group_df['PLAYER NAME'].values:
                    return True
    return False

def read_excel(file_path, sheet_name=None):
    return pd.read_excel(file_path, sheet_name=sheet_name)

def shuffle_and_sort_players(players_df):
    shuffled_players = players_df.sample(frac=1).reset_index(drop=True)
    shuffled_players.sort_values(by='BUGGY', ascending=True, inplace=True)
    return shuffled_players

def create_groups(shuffled_players, max_group_size=4, past_games_data=None, max_games=8):
    # Ask the user for the number of retries
    max_retries = simpledialog.askinteger(
        "Effort Level",
        "How hard do you want your computer to work to find new playing partners?\n"
        "Choose a scale from 1 to 100, where 100 means the maximum effort (100 retries).\n"
        "Note: Higher values may slow down your computer.",
        minvalue=1, maxvalue=100, initialvalue=10
    )

    # If the user cancels or closes the dialog, use a default value
    if max_retries is None:
        max_retries = 10

    best_grouping = None
    most_nones = -1  # Initialize with a low number

    if max_retries is None:
        max_retries = 10  # Default value if the user cancels the dialog


    # If the user cancels or closes the dialog, use a default value
    if max_retries is None:
        max_retries = 10

    best_grouping = None
    most_nones = -1  # Initialize with a low number


    for _ in range(max_retries):
        shuffled_players = shuffle_and_sort_players(shuffled_players)
        groups, played_together_pairs = _form_groups(shuffled_players, max_group_size, past_games_data, max_games)
        
        # Sorting groups with the logic mentioned before
        sorted_groups = sorted(groups, key=lambda x: (len(x) != 3, any(player[2] == 'YES' for player in x)))

        # Recreate the played_together_pairs based on sorted_groups
        played_together_pairs = []
        for group in sorted_groups:
            played_before_pairs = []
            for i in range(len(group)):
                for j in range(i + 1, len(group)):
                    if have_played_together(group[i][0], group[j][0], past_games_data, max_games):
                        played_before_pairs.append((group[i][0], group[j][0]))
            played_together_pairs.append(played_before_pairs)
        
        # Count 'None' entries which indicate no previous play together
        none_count = sum(1 for pairs in played_together_pairs if len(pairs) == 0)

        # If this is the best grouping so far, remember it
        if none_count > most_nones:
            most_nones = none_count
            best_grouping = sorted_groups  # Use the sorted groups
            best_played_together = played_together_pairs

    # Return the best grouping and played together pairs found during retries
    return best_grouping, best_played_together


    # If we never found a grouping, just return the last attempted grouping
    if best_grouping is None:
        return groups, played_together_pairs
    else:
        return best_grouping, best_played_together


def _form_groups(shuffled_players, max_group_size, past_games_data, max_games):
    buggy_count = sum(1 for _, player in shuffled_players.iterrows() if player['BUGGY'] == 'YES')
    total_players = len(shuffled_players)
    buggy_groups = (buggy_count + max_group_size - 1) // max_group_size
    remaining_players = total_players - buggy_groups * max_group_size
    total_groups_of_four = remaining_players // max_group_size
    remaining_players %= max_group_size

    if remaining_players > 0:
        groups_of_three = 1 + (max_group_size - 1 - remaining_players)
        total_groups_of_four -= groups_of_three
    else:
        groups_of_three = 0

    total_groups = total_groups_of_four + groups_of_three + buggy_groups
    groups = []
    group = []
    group_number = 1
    current_group_size = max_group_size if group_number <= buggy_groups else (3 if group_number <= buggy_groups + groups_of_three else 4)

    for _, player in shuffled_players.iterrows():
        group.append([player['PLAYER NAME'], player['HANDICAP'], 'YES' if player['BUGGY'] == 'YES' else ''])

        if len(group) == current_group_size:
            groups.append(group)
            group = []
            group_number += 1
            current_group_size = max_group_size if group_number <= buggy_groups else (3 if group_number <= buggy_groups + groups_of_three else 4)

    if group:
        groups.append(group)

 # Track pairs of players who have played together before
    played_together_pairs = []

    for group in groups:
        played_before_pairs = []
        for i in range(len(group)):
            for j in range(i + 1, len(group)):
                if have_played_together(group[i][0], group[j][0], past_games_data, max_games):
                    played_before_pairs.append((group[i][0], group[j][0]))

        played_together_pairs.append(played_before_pairs)

    return groups, played_together_pairs

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.borders import Border, Side
from pandas import read_excel, DataFrame
from datetime import datetime


def format_groups_sheet(ws, groups, played_together_pairs):
    # Set the headers
    headers = ['GROUP NUMBER', 'PLAYER NAME', 'HANDICAP', 'BUGGY', 'STABLEFORD', 'GROUP AVERAGE', 'NUM', 'PLAYED TOGETHER WITH']

    # Apply header formatting: bold, centered horizontally and vertically, font size 12
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment

    # Start writing group data from the second row
    row_num = 2

    for group_index, (group, pairs) in enumerate(zip(groups, played_together_pairs), start=1):
        for player in group:
            ws.cell(row=row_num, column=1, value=f"Group {group_index}")
            ws.cell(row=row_num, column=2, value=player[0])
            ws.cell(row=row_num, column=3, value=player[1])
            buggy_cell = ws.cell(row=row_num, column=4, value=player[2])
            # Leave STABLEFORD and GROUP AVERAGE blank for manual entry
            ws.cell(row=row_num, column=5, value='')
            ws.cell(row=row_num, column=6, value='')
            # Determine players who have played together
            played_before_with = ', '.join([p[1] for p in pairs if p[0] == player[0]] + [p[0] for p in pairs if p[1] == player[0]])
            played_before_with = played_before_with if played_before_with else 'None'
            ws.cell(row=row_num, column=8, value=played_before_with)
            row_num += 1
        # Leave a blank row between groups
        row_num += 1

    # Calculate the total number of players across all groups
    total_players = sum(len(group) for group in groups)

    # Write the 'NUMBER OF PLAYERS' total only once on the whole sheet
    ws.cell(row=row_num, column=7, value=total_players)
    ws.cell(row=row_num, column=7).font = Font(bold=True)  # Make it bold

    # Auto-adjust column widths to the widest text
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Apply all the remaining formatting for the entire used range
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=12)

    # Generate a unique table name based on the current date and time
    table_name = f"GroupData_{datetime.now().strftime('%Y%m%d%H%M%S')}"

    # Create a table for better data management
    table = Table(displayName=table_name, ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)
   
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook
import pandas as pd
from datetime import datetime

def main():
    # Initialize Tkinter
    root = tk.Tk()
    root.withdraw()  # Hide the main Tkinter window

    # Open file dialog with a message for the user
    input_file_path = filedialog.askopenfilename(
        title="Please select the Excel file Kevin", 
        filetypes=[("Excel files", "*.xlsx")]
    )
    
    # Check if a file was selected
    if not input_file_path:
        print("No file selected.")
        sys.exit(1)


    try:
        book = load_workbook(input_file_path)
    except FileNotFoundError:
        book = Workbook()

    # Read the players' data from the Excel sheet
    players_df = pd.read_excel(input_file_path, 'PLAYERS')
    players_df['BUGGY'] = players_df['BUGGY'].apply(lambda x: 'YES' if x == 'YES' else '')
    shuffled_players = shuffle_and_sort_players(players_df)
    past_games_data = {}

    # Load past game data
    for sheet_name in book.sheetnames:
        if sheet_name.startswith("Game "):
            past_game_df = pd.read_excel(input_file_path, sheet_name)
            past_game_df['PLAYER NAME'] = past_game_df['PLAYER NAME'].str.strip()
            past_game_df['GROUP NUMBER'] = past_game_df['GROUP NUMBER'].astype(str)
            past_games_data[sheet_name] = past_game_df

    max_games = 8
    groups, played_together = create_groups(shuffled_players, past_games_data=past_games_data, max_games=max_games)

    # Find the highest existing game number
    current_max_game_number = 0
    for sheet_name in book.sheetnames:
        if sheet_name.startswith("Game "):
            try:
                game_number = int(sheet_name.split(" ")[1])
                current_max_game_number = max(current_max_game_number, game_number)
            except ValueError:
                continue

    # Generate a new game number
    new_game_number = current_max_game_number + 1
    unique_sheet_name = f"Game {new_game_number}"
    new_sheet = book.create_sheet(title=unique_sheet_name)
    format_groups_sheet(new_sheet, groups, played_together)
    book.save(input_file_path)

if __name__ == "__main__":
    main()


